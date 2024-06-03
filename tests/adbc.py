# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
#
#
# def find_blue_cells(file_path, sheet_name):
#     # Load workbook
#     wb = load_workbook(file_path)
#     sheet = wb[sheet_name]
#
#     blue_cells = {}  # Dictionary to store the results
#
#     blue_found = False
#     start_row = None
#
#     # Loop through rows
#     for row in sheet.iter_rows(min_row=10, max_col=1):
#         cell = row[0]  # First cell in the row
#         print(cell.fill.start_color.index)
#         # Check cell fill color
#         if cell.fill and cell.fill.start_color.index == 'FF00FF00':
#             # Blue color found
#             if not blue_found:
#                 start_row = cell.row
#                 print("start_row: ", start_row)
#                 blue_found = True
#         elif blue_found:
#             # Blue color no longer found, record the range
#             blue_cells[cell.value] = cell.row - start_row + 1
#             blue_found = False
#
#     return blue_cells
#
#
# # Example usage
# file_path = 'Book2.xlsx'
# sheet_name = 'Sheet1'
# blue_cells_dict = find_blue_cells(file_path, sheet_name)
# print(blue_cells_dict)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def find_blue_cells(file_path, sheet_name, column_number):
    # Load workbook
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    blue_cells = []  # List to store the results

    # Loop through rows
    for row_idx in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=column_number)

        # Check cell fill color
        if cell.fill and cell.fill.start_color.index == 'FF00FF00':
            # Blue color found
            blue_cells.append((cell.value, cell.row))

    return blue_cells


# Example usage
file_path = 'Book2.xlsx'
sheet_name = 'Sheet1'
column_number = 1
blue_cells_list = find_blue_cells(file_path, sheet_name, column_number)
print(blue_cells_list)
