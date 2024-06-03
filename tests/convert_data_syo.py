import pandas as pd
from openpyxl import load_workbook


def find_color_cell(file_path, sheet_name, column_number):
    # Load workbook
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    gray_cells = []
    blue_cells = []

    # Loop through rows
    for row_idx in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=column_number)

        # Check cell fill color
        if cell.fill and cell.fill.start_color.index[2:] == '535353':
            # Blue color found
            gray_cells.append((cell.value, cell.row))
        elif cell.fill and cell.fill.start_color.index[2:] == '245269':
            blue_cells.append((cell.value, cell.row))
    return gray_cells, blue_cells


def dataframe_convert(file_path):
    sheet_name = 'Sheet1'
    column_number = 1
    df = pd.read_excel(file_path)
    gray_cells_list, blue_cells_list = find_color_cell(file_path, sheet_name, column_number)
    gray_cells_list = [('UNKNOW_device_group', 0)] + gray_cells_list + [('xxx', blue_cells_list[-1][1] - 3)]
    blue_cells_list = [('UNKNOW_device', 0)] + blue_cells_list
    new_column_names = ['comment_1', 'comment_2', 'comment_3', 'comment_4', 'comment_5']
    df.columns.values[-5:] = new_column_names

    index_gray_old = 0
    value_gray_old = 'xxx'
    for item in gray_cells_list:
        if item[0] == 'UNKNOW_device_group':
            value_gray_old = item[0]
            index_gray_old = item[1]
        else:
            df.loc[index_gray_old:item[1], 'device_group'] = value_gray_old
            value_gray_old = item[0]
            index_gray_old = item[1] - 2
    index_blue_old = 0
    value_blue_old = 'xxx'
    for item in blue_cells_list:
        if item[0] == 'UNKNOW_device':
            value_blue_old = item[0]
            index_blue_old = item[1] - 1
        else:
            df.loc[index_blue_old:item[1], 'device_name'] = value_blue_old
            value_blue_old = item[0]
            index_blue_old = item[1] - 2

    opt_index = df.index[df['CADICS ID'] == 'OptionCode'].values[0]
    # print(df.iloc[:opt_index - 1])
    df_end = df.iloc[:opt_index - 1]
    # df_end.to_excel('mandan222.xlsx', index=None)
    df_cleaned = df_end.dropna(subset=['Gr', 'CADICS ID'], how='all')
    # df_cleaned.to_excel("mandan_11.xlsx", index=None)
    df_end = df_end[['device_group', 'device_name', 'Gr', 'CADICS ID']]
    tuple_list = list(df_end.itertuples(index=False, name=None))
    return tuple_list


if __name__ == "__main__":
    file_path = r"C:\Users\KNT21617\Downloads\newken\仕様表\仕様表_PZ1K_copy - Copy.xlsx"
    tuple_list = dataframe_convert(file_path)
    print(tuple_list)
    # df.to_excel('output.xlsx', index=None)
