import pandas as pd
from openpyxl import load_workbook

# Open the Excel file
file_path = r"C:\Users\KNT21617\Downloads\New folder (4)\abc.xlsx"

# Read data from the Excel file
df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', header=None)
print(df)

# Search for the value 'linghting' in column D
search_value = 'linghting'
col_index = 3  # Column D
if search_value in df.iloc[:, col_index].values:
    # Get the position of the cell containing the value 'linghting'
    row_index = df[df.iloc[:, col_index] == search_value].index[0] + 1  # +1 because Excel rows start at 1
    col_letter = 'D'

    # Create a workbook and get the sheet
    wb = load_workbook(file_path)
    ws = wb['Sheet1']

    # Get the cell to check the background color
    cell = ws[f'{col_letter}{row_index}']

    # Get the background color of the cell
    fill = cell.fill
    start_color = fill.start_color.index  # Color code of the cell

    # Print the background color of the cell
    print(
        f"Found the value 'linghting' at cell {col_letter}{row_index}. The background color of this cell is: {start_color}.")
else:
    print("Did not find the value 'linghting' in column D.")
