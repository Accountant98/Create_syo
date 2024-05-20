import pandas as pd
from openpyxl import load_workbook

"""fuction:find_color
-Description: Find keywords in column D of excel and return the background color of that cell
-Input: df_ (dataframe), search_value_(string), file_path_(string)
-Output: background color of cell
"""


def find_color(df_, search_value_, file_path_):
    col_index = 3
    if search_value_ in df_.iloc[:, col_index].values:
        row_index = df_[df_.iloc[:, col_index] == search_value_].index[0] + 1
        col_letter = 'D'

        wb = load_workbook(file_path_)
        ws = wb['Sheet1']
        cell = ws[f'{col_letter}{row_index}']
        fill = cell.fill
        start_color = fill.start_color.index
        return start_color
    else:
        return ''


file_path = r'../../data/raw/input.xlsx'
df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', header=None)
search_value = 'linghting'
color = find_color(df, search_value, file_path)
print(color)
